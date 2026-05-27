"""
QBO Profit & Loss (standard P&L) — parser + Flask endpoint.

Make.com POSTs the .xlsx as multipart/form-data (field name: 'file')
to /process, this returns JSON, Make writes a row to Google Sheets.
One QBO P&L export → one row in the sheet.
"""

from __future__ import annotations

import calendar
import logging
import os
import re
from datetime import datetime
from io import BytesIO
from typing import Optional

from flask import Flask, jsonify, request
from openpyxl import load_workbook

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

app = Flask(__name__)

# Optional shared-secret auth. Set PIPELINE_API_KEY in Railway env vars.
# Make.com sends it in the 'X-Api-Key' header. Leave unset to skip auth.
API_KEY = os.environ.get("PIPELINE_API_KEY")


# ---------- Flask routes ----------

@app.route("/process", methods=["POST"])
def process():
    if API_KEY and request.headers.get("X-Api-Key") != API_KEY:
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    upload = request.files.get("file")
    if upload is None:
        return jsonify({"ok": False, "error": "missing file"}), 400

    file_bytes = upload.read()
    if not file_bytes:
        return jsonify({"ok": False, "error": "empty file"}), 400

    try:
        parsed = parse(file_bytes)
    except ValueError as e:
        log.warning("parse failed: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 422
    except Exception as e:  # noqa: BLE001
        log.exception("unexpected parse error")
        return jsonify({"ok": False, "error": f"parse failed: {e}"}), 500

    log.info("parsed report_date=%s fields=%d",
             parsed.get("report_date"), len(parsed) - 1)
    return jsonify(parsed), 200


@app.route("/", methods=["GET"])
def health():
    return jsonify({"ok": True, "service": "pl-parser"}), 200


# ---------- Parser ----------

def parse(file_bytes: bytes) -> dict:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    report_date = _find_report_end_date(rows)
    if report_date is None:
        raise ValueError(
            "Could not find report date range in header. "
            "Is this actually a QBO Profit & Loss export?"
        )

    result: dict = {"report_date": report_date}

    for row in rows:
        if not row or len(row) < 2:
            continue
        label, value = row[0], row[1]
        if label is None or value is None:
            continue
        label_str = str(label).strip()
        if label_str in ("Total", ""):
            continue
        if isinstance(value, (int, float)):
            result[label_str] = round(float(value), 2)

    return result


_MONTHS = (r"(?:January|February|March|April|May|June|"
           r"July|August|September|October|November|December)")


def _find_report_end_date(rows: list) -> Optional[str]:
    """Pull end date from QBO header, handling multiple formats:
        'May 1, 2025-May 27, 2026'        -> 05/27/2026
        'January 1 - December 31, 2025'   -> 12/31/2025
        'January 1-31, 2025'              -> 01/31/2025
        'January 2026'                    -> 01/31/2026
        '2025'                            -> 12/31/2025
        '1/1/2025-12/31/2025'             -> 12/31/2025
    """
    for i in range(min(10, len(rows))):
        if not rows[i] or rows[i][0] is None:
            continue
        text = str(rows[i][0]).strip()
        result = _parse_period_string(text)
        if result:
            return result
    return None


def _parse_period_string(text: str) -> Optional[str]:
    # 1. "...-Month Day, Year" (any range ending with full month/day/year)
    m = re.search(rf"-\s*({_MONTHS})\s+(\d+),?\s*(\d{{4}})\s*$", text)
    if m:
        try:
            return datetime.strptime(
                f"{m.group(1)} {m.group(2)} {m.group(3)}", "%B %d %Y"
            ).strftime("%m/%d/%Y")
        except ValueError:
            pass

    # 2. "Month Day-Day, Year" (same-month range like "January 1-31, 2025")
    m = re.search(
        rf"({_MONTHS})\s+\d+\s*-\s*(\d+),?\s*(\d{{4}})\s*$", text
    )
    if m:
        try:
            return datetime.strptime(
                f"{m.group(1)} {m.group(2)} {m.group(3)}", "%B %d %Y"
            ).strftime("%m/%d/%Y")
        except ValueError:
            pass

    # 3. "Month Year" (single month like "January 2026") -> last day of month
    m = re.match(rf"^({_MONTHS})\s+(\d{{4}})\s*$", text)
    if m:
        try:
            dt = datetime.strptime(f"{m.group(1)} 1 {m.group(2)}", "%B %d %Y")
            last_day = calendar.monthrange(dt.year, dt.month)[1]
            return dt.replace(day=last_day).strftime("%m/%d/%Y")
        except ValueError:
            pass

    # 4. "MM/DD/YYYY-MM/DD/YYYY" (numeric range) -> end date
    m = re.search(r"-\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$", text)
    if m:
        try:
            return datetime.strptime(
                f"{m.group(1)}/{m.group(2)}/{m.group(3)}", "%m/%d/%Y"
            ).strftime("%m/%d/%Y")
        except ValueError:
            pass

    # 5. "YYYY" alone -> Dec 31 of that year
    m = re.match(r"^(\d{4})\s*$", text)
    if m:
        return f"12/31/{m.group(1)}"

    return None


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
